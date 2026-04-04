#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
import time
from collections import deque
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Callable
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from PySide6.QtCore import QObject, QDate, QThread, Signal, Slot
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QSpinBox,
    QDoubleSpinBox,
    QVBoxLayout,
    QWidget,
    QGridLayout,
    QDateEdit,
)


LOGIN_URL = "https://dealer.parts-unlimited.com/api/login"
ORDERS_URL = "https://dealer.parts-unlimited.com/api/orders/submitted"
DEFAULT_DETAIL_URL_TEMPLATE = "https://dealer.parts-unlimited.com/api/orders/{order_id}"
AUTHORIZED_USERS_URL = "https://raw.githubusercontent.com/rmtodd2/Lightning/main/Number.txt"

TIMEOUT = 30
PAGE_SIZE = 100


class RateLimiter:
    def __init__(
        self,
        max_requests: int = 900,
        window_seconds: int = 300,
        debug: bool = False,
        log: Optional[Callable[[str], None]] = None,
    ) -> None:
        self.max_requests = max_requests
        self.window_seconds = window_seconds
        self.debug = debug
        self.log = log
        self.request_times: deque[float] = deque()

    def wait_for_slot(self) -> None:
        now = time.time()

        while self.request_times and now - self.request_times[0] >= self.window_seconds:
            self.request_times.popleft()

        if len(self.request_times) >= self.max_requests:
            sleep_for = self.window_seconds - (now - self.request_times[0]) + 1.0
            if sleep_for > 0:
                if self.debug and self.log:
                    self.log(
                        f"Rate limit guard: {len(self.request_times)} requests in the last "
                        f"{self.window_seconds} seconds. Sleeping for {sleep_for:.1f} seconds..."
                    )
                time.sleep(sleep_for)

        now = time.time()
        while self.request_times and now - self.request_times[0] >= self.window_seconds:
            self.request_times.popleft()

        self.request_times.append(now)


def parse_iso_datetime(value: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def parse_order_date(value: str) -> Optional[date]:
    dt = parse_iso_datetime(value)
    return dt.date() if dt else None


def mask_value(value: str, prefix: int = 2, suffix: int = 2) -> str:
    if not value:
        return "<missing>"
    if len(value) <= prefix + suffix:
        return "*" * len(value)
    middle = "*" * max(4, len(value) - prefix - suffix)
    return f"{value[:prefix]}{middle}{value[-suffix:]}"


def get_tracking_numbers(item: Dict[str, Any]) -> List[str]:
    cartons = item.get("shippingCartons") or []
    values: List[str] = []
    for carton in cartons:
        tracking = str(carton.get("trackingNumber", "") or "").strip()
        if tracking:
            values.append(tracking)
    return sorted(set(values))


def get_quantity_shipped(item: Dict[str, Any]) -> int:
    value = item.get("quantityShipped", 0)
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def get_quantity(item: Dict[str, Any]) -> int:
    value = item.get("quantity", 0)
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def order_number(order: Dict[str, Any]) -> str:
    return str(order.get("mfOrderNumber", "") or "").strip()


def submitted_date(order: Dict[str, Any]) -> str:
    dt = parse_iso_datetime(str(order.get("submittedAt", "") or ""))
    return dt.date().isoformat() if dt else ""


def request_with_retry(
    session: requests.Session,
    method: str,
    url: str,
    *,
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    debug: bool = False,
    log: Optional[Callable[[str], None]] = None,
    **kwargs: Any,
) -> requests.Response:
    retryable_statuses = {429, 500, 502, 503, 504}
    last_exception: Optional[Exception] = None

    for attempt in range(1, 7):
        rate_limiter.wait_for_slot()

        try:
            response = session.request(method, url, timeout=TIMEOUT, **kwargs)

            if response.status_code in retryable_statuses:
                retry_after = response.headers.get("Retry-After")
                if retry_after:
                    try:
                        wait_time = float(retry_after)
                    except ValueError:
                        wait_time = backoff_seconds * attempt
                else:
                    wait_time = backoff_seconds * attempt

                if log:
                    log(
                        f"HTTP {response.status_code} on {method} {url}. "
                        f"Retrying in {wait_time:.1f} seconds (attempt {attempt}/6)..."
                    )

                time.sleep(wait_time)
                continue

            response.raise_for_status()

            if request_delay > 0:
                time.sleep(request_delay)

            return response

        except requests.RequestException as exc:
            last_exception = exc
            wait_time = backoff_seconds * attempt
            if log:
                log(
                    f"Request error on {method} {url}: {exc}. "
                    f"Retrying in {wait_time:.1f} seconds (attempt {attempt}/6)..."
                )
            time.sleep(wait_time)

    if last_exception:
        raise last_exception

    raise RuntimeError(f"Request failed unexpectedly: {method} {url}")


def extract_username_from_email(email: str) -> str:
    value = str(email or "").strip()
    if not value or "@" not in value:
        return ""
    return value.split("@", 1)[0].strip()


def normalize_github_raw_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.netloc.lower() in {"raw.githubusercontent.com"}:
        return url

    if parsed.netloc.lower() in {"github.com", "www.github.com"}:
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) >= 5 and parts[2] == "blob":
            owner, repo, _, branch = parts[:4]
            file_path = "/".join(parts[4:])
            return f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{file_path}"

    return url


def fetch_authorized_users(
    session: requests.Session,
    authorized_users_url: str,
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    debug: bool,
    log: Callable[[str], None],
) -> Set[str]:
    raw_url = normalize_github_raw_url(authorized_users_url)

    if debug:
        log(f"Authorized users URL: {raw_url}")

    response = request_with_retry(
        session,
        "GET",
        raw_url,
        rate_limiter=rate_limiter,
        request_delay=request_delay,
        backoff_seconds=backoff_seconds,
        debug=debug,
        log=log,
    )
    response.raise_for_status()

    users: Set[str] = set()
    for line in response.text.splitlines():
        name = line.strip().replace("\ufeff", "")
        if name:
            users.add(name)

    return users


def login(
    session: requests.Session,
    dealer_code: str,
    username: str,
    password: str,
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    debug: bool,
    log: Callable[[str], None],
) -> Dict[str, Any]:
    payload = {
        "dealerCode": dealer_code,
        "username": username,
        "password": password,
    }

    if debug:
        log(f"Login URL: {LOGIN_URL}")
        log(
            "Login payload: "
            + str(
                {
                    "dealerCode": dealer_code,
                    "username": mask_value(username),
                    "password": "<masked>",
                }
            )
        )

    response = request_with_retry(
        session,
        "PUT",
        LOGIN_URL,
        json=payload,
        rate_limiter=rate_limiter,
        request_delay=request_delay,
        backoff_seconds=backoff_seconds,
        debug=debug,
        log=log,
    )
    response.raise_for_status()
    return response.json()


def authorize_logged_in_user(
    session: requests.Session,
    login_response: Dict[str, Any],
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    debug: bool,
    log: Callable[[str], None],
) -> str:
    email = str(login_response.get("email", "") or "").strip()
    username_from_email = extract_username_from_email(email)

    if not username_from_email or not re.fullmatch(r"\d+", username_from_email):
        raise PermissionError("Access Denied, This user is not authorized to use this app.")

    authorized_users = fetch_authorized_users(
        session=session,
        authorized_users_url=AUTHORIZED_USERS_URL,
        rate_limiter=rate_limiter,
        request_delay=request_delay,
        backoff_seconds=backoff_seconds,
        debug=debug,
        log=log,
    )

    if username_from_email not in authorized_users:
        raise PermissionError("Access Denied, This user is not authorized to use this app.")

    return username_from_email


def fetch_submitted_orders_page(
    session: requests.Session,
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    offset: int,
    limit: int,
    debug: bool,
    log: Callable[[str], None],
) -> Dict[str, Any]:
    params = {
        "offset": offset,
        "limit": limit,
        "t": int(time.time() * 1000),
    }

    if debug:
        log(f"Orders URL: {ORDERS_URL}")
        log(f"Orders params: {params}")

    response = request_with_retry(
        session,
        "GET",
        ORDERS_URL,
        params=params,
        rate_limiter=rate_limiter,
        request_delay=request_delay,
        backoff_seconds=backoff_seconds,
        debug=debug,
        log=log,
    )
    return response.json()


def fetch_order_detail(
    session: requests.Session,
    order: Dict[str, Any],
    detail_url_template: str,
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    debug: bool,
    log: Callable[[str], None],
) -> Dict[str, Any]:
    order_id = str(order.get("id", "") or "").strip()
    mf_order_number = order_number(order)

    if not order_id:
        raise ValueError(f"Order {mf_order_number or '<unknown>'} has no id field.")

    url = detail_url_template.format(
        order_id=order_id,
        order_number=mf_order_number,
    )

    if debug:
        log(f"Fetching detail for order {mf_order_number} from {url}")

    response = request_with_retry(
        session,
        "GET",
        url,
        rate_limiter=rate_limiter,
        request_delay=request_delay,
        backoff_seconds=backoff_seconds,
        debug=debug,
        log=log,
    )
    return response.json()


def build_invoice_quantity_by_line_id(order_detail: Dict[str, Any]) -> Dict[str, int]:
    quantities: Dict[str, int] = {}
    invoices = order_detail.get("invoices") or []

    for invoice in invoices:
        for item in invoice.get("items") or []:
            line_id = str(item.get("id", "") or "").strip()
            if not line_id:
                continue
            quantities[line_id] = quantities.get(line_id, 0) + get_quantity_shipped(item)

    return quantities


def find_tracked_uninvoiced_lines(order_detail: Dict[str, Any]) -> List[Dict[str, Any]]:
    invoice_qty_by_line_id = build_invoice_quantity_by_line_id(order_detail)
    results: List[Dict[str, Any]] = []

    for item in order_detail.get("items") or []:
        tracking_numbers = get_tracking_numbers(item)
        if not tracking_numbers:
            continue

        shipped_qty = get_quantity_shipped(item)
        if shipped_qty <= 0:
            continue

        line_id = str(item.get("id", "") or "").strip()
        invoiced_qty = invoice_qty_by_line_id.get(line_id, 0)

        if invoiced_qty >= shipped_qty:
            continue

        part = item.get("part") or {}
        result = {
            "dealer_number": str(order_detail.get("dealerCode", "") or ""),
            "order_number": str(order_detail.get("mfOrderNumber", "") or ""),
            "order_id": str(order_detail.get("id", "") or ""),
            "order_status": str(order_detail.get("status", "") or ""),
            "submitted_at": str(order_detail.get("submittedAt", "") or ""),
            "submitted_date": submitted_date(order_detail),
            "line_id": line_id,
            "part_number": str(item.get("partNumber", "") or ""),
            "punctuated_part_number": str(part.get("punctuatedPartNumber", "") or ""),
            "description": str(part.get("description", "") or ""),
            "brand": str(((part.get("brand") or {}).get("name", "")) or ""),
            "line_status": str(item.get("lineStatus", "") or ""),
            "quantity": get_quantity(item),
            "quantity_shipped": shipped_qty,
            "quantity_invoiced": invoiced_qty,
            "tracking_numbers": ", ".join(tracking_numbers),
            "ship_date": str(item.get("shipDate", "") or ""),
            "ship_via": str(item.get("shipVia", "") or ""),
            "location_code": str(item.get("locationCode", "") or ""),
        }
        results.append(result)

    return results



def summarize_row(row: Dict[str, Any], index: int) -> str:
    return (
        f"[{index}] Dealer #: {row['dealer_number']} | "
        f"Order Number: {row['order_number']} | "
        f"Part: {row['punctuated_part_number'] or row['part_number']} | "
        f"Tracking: {row['tracking_numbers']} | "
        f"Shipped: {row['quantity_shipped']} | "
        f"Invoiced: {row['quantity_invoiced']}"
    )



def clean_prefixed_value(value: str, label: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^\[\d+\]\s*", "", text)
    prefix = f"{label}:"
    if text.upper().startswith(prefix.upper()):
        text = text[len(prefix):].strip()
    return text



def build_summary_rows(rows: List[Dict[str, Any]]) -> List[Tuple[str, str]]:
    seen: Set[Tuple[str, str]] = set()
    summary_rows: List[Tuple[str, str]] = []

    for row in rows:
        dealer = clean_prefixed_value(row.get("dealer_number", ""), "Dealer #")
        order = clean_prefixed_value(row.get("order_number", ""), "Order Number")

        key = (dealer.upper(), order.upper())
        if key in seen:
            continue

        seen.add(key)
        summary_rows.append((dealer, order))

    return summary_rows



def autofit_worksheet_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        adjusted_width = min(max(max_length + 2, 10), 80)
        ws.column_dimensions[column_letter].width = adjusted_width



def write_excel_output(path: Path, full_rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = ["Dealer #", "Order Number"]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    for dealer, order in build_summary_rows(full_rows):
        ws_summary.append([dealer, order])

    autofit_worksheet_columns(ws_summary)

    ws_full = wb.create_sheet("Full Output")
    full_headers = [
        "Dealer #",
        "Order Number",
        "Order ID",
        "Order Status",
        "Submitted Date",
        "Submitted At",
        "Line ID",
        "Part",
        "Part Number Raw",
        "Description",
        "Brand",
        "Line Status",
        "Quantity",
        "Shipped",
        "Invoiced",
        "Tracking",
        "Ship Date",
        "Ship Via",
        "Location",
    ]
    ws_full.append(full_headers)
    for cell in ws_full[1]:
        cell.font = Font(bold=True)

    for row in full_rows:
        ws_full.append(
            [
                clean_prefixed_value(row.get("dealer_number", ""), "Dealer #"),
                clean_prefixed_value(row.get("order_number", ""), "Order Number"),
                row.get("order_id", ""),
                row.get("order_status", ""),
                row.get("submitted_date", ""),
                row.get("submitted_at", ""),
                row.get("line_id", ""),
                row.get("punctuated_part_number", "") or row.get("part_number", ""),
                row.get("part_number", ""),
                row.get("description", ""),
                row.get("brand", ""),
                row.get("line_status", ""),
                row.get("quantity", 0),
                row.get("quantity_shipped", 0),
                row.get("quantity_invoiced", 0),
                row.get("tracking_numbers", ""),
                row.get("ship_date", ""),
                row.get("ship_via", ""),
                row.get("location_code", ""),
            ]
        )

    autofit_worksheet_columns(ws_full)
    wb.save(path)



def fetch_all_requested_pages(
    session: requests.Session,
    detail_url_template: str,
    oldest_date: date,
    newest_date: date,
    rate_limiter: RateLimiter,
    request_delay: float,
    backoff_seconds: float,
    debug: bool,
    log: Callable[[str], None],
) -> Dict[str, Any]:
    matched_rows: List[Dict[str, Any]] = []
    scanned_orders = 0
    detailed_orders = 0
    offset = 0
    page_number = 0
    total_count: Optional[int] = None
    stop_reason = ""

    while True:
        page_number += 1
        page = fetch_submitted_orders_page(
            session=session,
            rate_limiter=rate_limiter,
            request_delay=request_delay,
            backoff_seconds=backoff_seconds,
            offset=offset,
            limit=PAGE_SIZE,
            debug=debug,
            log=log,
        )

        page_orders = page.get("orders", []) or []
        total_count = int(page.get("totalCount", 0) or 0)
        page_limit = int(page.get("limit", PAGE_SIZE) or PAGE_SIZE)

        if not page_orders:
            log(f"Scanning page {page_number}... no orders returned.")
            break

        first_page_date = parse_order_date(str(page_orders[0].get("submittedAt", "") or ""))
        if first_page_date:
            log(
                f"Scanning page {page_number} "
                f"(offset {offset}, page starts with {first_page_date.isoformat()})..."
            )
        else:
            log(f"Scanning page {page_number} (offset {offset})...")

        should_stop_after_page = False

        for order in page_orders:
            order_date = parse_order_date(str(order.get("submittedAt", "") or ""))
            if order_date and order_date < oldest_date:
                should_stop_after_page = True
                stop_reason = (
                    f"Reached order date {order_date.isoformat()}, which is older than the "
                    f"requested oldest date {oldest_date.isoformat()}."
                )
                break

            if order_date and order_date > newest_date:
                continue

            scanned_orders += 1
            num = order_number(order)

            order_status = str(order.get("status", "") or "").upper().strip()
            if order_status != "SHIPPED":
                if debug:
                    log(
                        f"Skipping order {num or order.get('id') or '<unknown>'} "
                        f"with status {order_status or '<blank>'}"
                    )
                continue

            detail = fetch_order_detail(
                session=session,
                order=order,
                detail_url_template=detail_url_template,
                rate_limiter=rate_limiter,
                request_delay=request_delay,
                backoff_seconds=backoff_seconds,
                debug=debug,
                log=log,
            )
            detailed_orders += 1
            matched_rows.extend(find_tracked_uninvoiced_lines(detail))

        if should_stop_after_page:
            break

        if len(page_orders) < page_limit:
            stop_reason = "Reached the last available page of orders."
            break

        offset += page_limit

    return {
        "rows": matched_rows,
        "scanned_orders": scanned_orders,
        "detailed_orders": detailed_orders,
        "pages_scanned": page_number,
        "total_count": total_count if total_count is not None else scanned_orders,
        "stop_reason": stop_reason,
        "oldest_date": oldest_date.isoformat(),
        "newest_date": newest_date.isoformat(),
    }


class ScanWorker(QObject):
    log_message = Signal(str)
    finished = Signal(bool, str)

    def __init__(
        self,
        dealer_code: str,
        username: str,
        password: str,
        oldest_date: date,
        newest_date: date,
        excel_file: str,
        request_delay: float,
        backoff_seconds: float,
        rate_limit_max_requests: int,
        rate_limit_window_seconds: int,
        debug: bool,
    ) -> None:
        super().__init__()
        self.dealer_code = dealer_code
        self.username = username
        self.password = password
        self.oldest_date = oldest_date
        self.newest_date = newest_date
        self.excel_file = excel_file
        self.request_delay = request_delay
        self.backoff_seconds = backoff_seconds
        self.rate_limit_max_requests = rate_limit_max_requests
        self.rate_limit_window_seconds = rate_limit_window_seconds
        self.debug = debug

    def log(self, message: str) -> None:
        self.log_message.emit(message)

    @Slot()
    def run(self) -> None:
        try:
            excel_path = Path(self.excel_file)
            excel_path.parent.mkdir(parents=True, exist_ok=True)

            rate_limiter = RateLimiter(
                max_requests=self.rate_limit_max_requests,
                window_seconds=self.rate_limit_window_seconds,
                debug=self.debug,
                log=self.log,
            )

            with requests.Session() as session:
                login_response = login(
                    session=session,
                    dealer_code=self.dealer_code,
                    username=self.username,
                    password=self.password,
                    rate_limiter=rate_limiter,
                    request_delay=self.request_delay,
                    backoff_seconds=self.backoff_seconds,
                    debug=self.debug,
                    log=self.log,
                )
                authorized_username = authorize_logged_in_user(
                    session=session,
                    login_response=login_response,
                    rate_limiter=rate_limiter,
                    request_delay=self.request_delay,
                    backoff_seconds=self.backoff_seconds,
                    debug=self.debug,
                    log=self.log,
                )
                data = fetch_all_requested_pages(
                    session=session,
                    detail_url_template=DEFAULT_DETAIL_URL_TEMPLATE,
                    oldest_date=self.oldest_date,
                    newest_date=self.newest_date,
                    rate_limiter=rate_limiter,
                    request_delay=self.request_delay,
                    backoff_seconds=self.backoff_seconds,
                    debug=self.debug,
                    log=self.log,
                )

            all_rows = data.get("rows", []) or []

            write_excel_output(excel_path, all_rows)

            self.log("")
            self.log(
                f"Date range searched: {data.get('oldest_date', '')} to {data.get('newest_date', '')}"
            )
            self.log(f"Pages scanned: {data.get('pages_scanned', 0)}")
            self.log(f"Orders scanned: {data.get('scanned_orders', 0)}")
            self.log(f"Order details fetched: {data.get('detailed_orders', 0)}")
            self.log(f"Tracked/uninvoiced lines found this scan: {len(all_rows)}")
            if data.get("stop_reason"):
                self.log(f"Stop reason: {data.get('stop_reason')}")
            self.log(f"Excel file: {excel_path}")
            self.log("")

            if all_rows:
                self.log("Tracked parts without invoice:")
                for idx, row in enumerate(all_rows, start=1):
                    self.log(summarize_row(row, idx))
            else:
                self.log("No tracked-but-uninvoiced parts found.")

            self.finished.emit(True, str(excel_path))

        except PermissionError as exc:
            self.log(str(exc))
            self.finished.emit(False, str(exc))
        except requests.HTTPError as exc:
            response = exc.response
            if response is not None:
                self.log(f"HTTP error: {response.status_code}")
                self.log(response.text)
            else:
                self.log(f"HTTP error: {exc}")
            self.finished.emit(False, "HTTP error")
        except requests.RequestException as exc:
            self.log(f"Network error: {exc}")
            self.finished.emit(False, "Network error")
        except Exception as exc:
            self.log(f"Unexpected error: {exc}")
            self.finished.emit(False, str(exc))


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Tracked Uninvoiced Parts Scanner")
        self.resize(980, 700)

        self.worker_thread: Optional[QThread] = None
        self.worker: Optional[ScanWorker] = None

        central = QWidget()
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)
        form_layout = QGridLayout()

        self.dealer_edit = QLineEdit()
        self.user_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)

        self.oldest_date_edit = QDateEdit()
        self.oldest_date_edit.setCalendarPopup(True)
        self.oldest_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.oldest_date_edit.setDate(QDate.currentDate().addDays(-7))

        self.newest_date_edit = QDateEdit()
        self.newest_date_edit.setCalendarPopup(True)
        self.newest_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.newest_date_edit.setDate(QDate.currentDate())

        self.output_edit = QLineEdit(str(Path.cwd() / "tracked_uninvoiced_parts.xlsx"))
        self.output_browse_btn = QPushButton("Browse...")
        self.output_browse_btn.clicked.connect(self.browse_output_file)

        self.request_delay_spin = QDoubleSpinBox()
        self.request_delay_spin.setRange(0.0, 10.0)
        self.request_delay_spin.setDecimals(2)
        self.request_delay_spin.setSingleStep(0.05)
        self.request_delay_spin.setValue(0.35)

        self.backoff_spin = QDoubleSpinBox()
        self.backoff_spin.setRange(0.1, 60.0)
        self.backoff_spin.setDecimals(1)
        self.backoff_spin.setSingleStep(0.5)
        self.backoff_spin.setValue(2.0)

        self.rate_limit_requests_spin = QSpinBox()
        self.rate_limit_requests_spin.setRange(1, 5000)
        self.rate_limit_requests_spin.setValue(900)

        self.rate_limit_window_spin = QSpinBox()
        self.rate_limit_window_spin.setRange(1, 3600)
        self.rate_limit_window_spin.setValue(300)

        form_layout.addWidget(QLabel("Dealer #:"), 0, 0)
        form_layout.addWidget(self.dealer_edit, 0, 1)
        form_layout.addWidget(QLabel("User ID:"), 1, 0)
        form_layout.addWidget(self.user_edit, 1, 1)
        form_layout.addWidget(QLabel("Password:"), 2, 0)
        form_layout.addWidget(self.password_edit, 2, 1)
        form_layout.addWidget(QLabel("Oldest date to search:"), 3, 0)
        form_layout.addWidget(self.oldest_date_edit, 3, 1)
        form_layout.addWidget(QLabel("Newest date to search:"), 4, 0)
        form_layout.addWidget(self.newest_date_edit, 4, 1)

        form_layout.addWidget(QLabel("Output file:"), 5, 0)
        output_row = QHBoxLayout()
        output_row.addWidget(self.output_edit, 1)
        output_row.addWidget(self.output_browse_btn)
        form_layout.addLayout(output_row, 5, 1)

        form_layout.addWidget(QLabel("Request delay (sec):"), 6, 0)
        form_layout.addWidget(self.request_delay_spin, 6, 1)
        form_layout.addWidget(QLabel("Retry backoff (sec):"), 7, 0)
        form_layout.addWidget(self.backoff_spin, 7, 1)
        form_layout.addWidget(QLabel("Rate limit max requests:"), 8, 0)
        form_layout.addWidget(self.rate_limit_requests_spin, 8, 1)
        form_layout.addWidget(QLabel("Rate limit window (sec):"), 9, 0)
        form_layout.addWidget(self.rate_limit_window_spin, 9, 1)

        main_layout.addLayout(form_layout)

        button_row = QHBoxLayout()
        self.start_btn = QPushButton("Start Scan")
        self.start_btn.clicked.connect(self.start_scan)
        button_row.addWidget(self.start_btn)
        button_row.addStretch(1)
        main_layout.addLayout(button_row)

        main_layout.addWidget(QLabel("Console Output:"))
        self.console = QPlainTextEdit()
        self.console.setReadOnly(True)
        self.console.setLineWrapMode(QPlainTextEdit.NoWrap)
        main_layout.addWidget(self.console, 1)

    def browse_output_file(self) -> None:
        current = self.output_edit.text().strip() or str(Path.cwd() / "tracked_uninvoiced_parts.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Choose Output Excel File",
            current,
            "Excel Files (*.xlsx)",
        )
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.output_edit.setText(path)

    def append_log(self, message: str) -> None:
        self.console.appendPlainText(message)
        sb = self.console.verticalScrollBar()
        sb.setValue(sb.maximum())

    def validate_inputs(self) -> bool:
        if not self.dealer_edit.text().strip():
            QMessageBox.warning(self, "Missing Value", "Please enter the Dealer #.")
            return False
        if not self.user_edit.text().strip():
            QMessageBox.warning(self, "Missing Value", "Please enter the User ID.")
            return False
        if not self.password_edit.text().strip():
            QMessageBox.warning(self, "Missing Value", "Please enter the Password.")
            return False
        if not self.output_edit.text().strip():
            QMessageBox.warning(self, "Missing Value", "Please choose an output file.")
            return False

        oldest = self.oldest_date_edit.date().toPython()
        newest = self.newest_date_edit.date().toPython()
        if oldest > newest:
            QMessageBox.warning(
                self,
                "Invalid Date Range",
                "The oldest date cannot be later than the newest date.",
            )
            return False
        return True

    def start_scan(self) -> None:
        if not self.validate_inputs():
            return

        output_path = Path(self.output_edit.text().strip())
        oldest_date = self.oldest_date_edit.date().toPython()
        newest_date = self.newest_date_edit.date().toPython()

        self.console.clear()
        self.append_log("Starting scan...")
        self.append_log(
            f"Searching from: {oldest_date.isoformat()} to {newest_date.isoformat()}"
        )
        self.append_log(f"Output file: {output_path}")
        self.append_log("")

        self.start_btn.setEnabled(False)
        self.output_browse_btn.setEnabled(False)

        self.worker_thread = QThread()
        self.worker = ScanWorker(
            dealer_code=self.dealer_edit.text().strip(),
            username=self.user_edit.text().strip(),
            password=self.password_edit.text().strip(),
            oldest_date=oldest_date,
            newest_date=newest_date,
            excel_file=str(output_path),
            request_delay=float(self.request_delay_spin.value()),
            backoff_seconds=float(self.backoff_spin.value()),
            rate_limit_max_requests=int(self.rate_limit_requests_spin.value()),
            rate_limit_window_seconds=int(self.rate_limit_window_spin.value()),
            debug=False,
        )
        self.worker.moveToThread(self.worker_thread)

        self.worker_thread.started.connect(self.worker.run)
        self.worker.log_message.connect(self.append_log)
        self.worker.finished.connect(self.on_scan_finished)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)

        self.worker_thread.start()

    @Slot(bool, str)
    def on_scan_finished(self, success: bool, result: str) -> None:
        self.start_btn.setEnabled(True)
        self.output_browse_btn.setEnabled(True)

        if success:
            self.append_log("")
            self.append_log("Scan completed successfully.")
            QMessageBox.information(
                self,
                "Complete",
                f"Scan completed successfully.\n\nWorkbook saved to:\n{result}",
            )
        else:
            self.append_log("")
            self.append_log("Scan failed.")
            QMessageBox.critical(
                self,
                "Scan Failed",
                f"The scan did not complete.\n\n{result}",
            )

        self.worker = None
        self.worker_thread = None


def main() -> int:
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
